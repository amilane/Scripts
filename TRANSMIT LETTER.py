# -*- coding: utf-8 -*-
import openpyxl
import os
from PyPDF2 import PdfFileReader


# Укажите имя папки выпуска, например 0055-P2-ARM-CPC-TRM-00000
number = '0055-P2-ARM-CPC-TRM-11111'





#===========================================================================================
pathToRelease = r'\\arena\ARMO-GROUP\ОБЪЕКТЫ\В_РАБОТЕ\41XX_AGPZ\04-ИсхДок\Выпуск\\' + number
pathToDocuments = pathToRelease + r'\Documents'

pathToVDR = r"\\arena\ARMO-GROUP\ОБЪЕКТЫ\В_РАБОТЕ\41XX_AGPZ\30-РД\02-ГИП\TRANSMITTAL VDR.xlsx"
wbVDR = openpyxl.load_workbook(pathToVDR, data_only=True)
vdrSheet = wbVDR.get_sheet_by_name('VDR')
vdrCsvSheet = wbVDR.get_sheet_by_name('VDR CSV')
vdrRows = range(12, 291)

pathToTemplate = r'\\arena\ARMO-GROUP\ОБЪЕКТЫ\В_РАБОТЕ\41XX_AGPZ\30-РД\02-ГИП\TRANSMITTAL TEMPLATE.xlsx'
wbTemplate = openpyxl.load_workbook(pathToTemplate, data_only=True)
templateSheet = wbTemplate.get_sheet_by_name('TRM Template 08-Feb-2018')

pathToTemplate_CSV = r'\\arena\ARMO-GROUP\ОБЪЕКТЫ\В_РАБОТЕ\41XX_AGPZ\30-РД\02-ГИП\TRANSMITTAL TEMPLATE_CSV.xlsx'
wbTemplate_CSV = openpyxl.load_workbook(pathToTemplate_CSV, data_only=True)
templateSheet_CSV = wbTemplate_CSV.get_sheet_by_name('Document Load')


# 29-fileName, 16-docClass, 25-disc, 27-docType, 28-docNo, 
# 30-nameDocEn, 31-nameDocRu, 34-issue, 35-Rev, 36-order, 41-date

vdrDict = {vdrSheet.cell(row = row, column = 29).value:\
            [vdrSheet.cell(row = row, column = 16).value,\
            vdrSheet.cell(row = row, column = 25).value,\
            vdrSheet.cell(row = row, column = 27).value,\
            vdrSheet.cell(row = row, column = 28).value,\
            vdrSheet.cell(row = row, column = 30).value,\
            vdrSheet.cell(row = row, column = 31).value,\
            vdrSheet.cell(row = row, column = 34).value,\
            vdrSheet.cell(row = row, column = 35).value,\
            vdrSheet.cell(row = row, column = 36).value,\
            vdrSheet.cell(row = row, column = 41).value\
            ] for row in vdrRows}

# 1-fileName, 2-SubprojectCode
# 3-Operation Complex (rus), 4-Operation Complex (eng)
# 5-Start-up Complex (rus), 6-Start-up Complex (eng)
# 7-Title Object (rus), 8-Title Object (eng)
# 9-Title Number (rus), 10-Title Number (eng)
# 11-Package Name (rus), 12-Package Name (eng)
# 13-Sequence number, 14-Document Class Code
# 15-Discipline (rus), 16-Discipline (eng)
# 17-Construction Contractor, 18-Construction Type

vdrCsvDict = {vdrCsvSheet.cell(row = row, column = 1).value:\
            [vdrCsvSheet.cell(row = row, column = 2).value,\
            vdrCsvSheet.cell(row = row, column = 3).value,\
            vdrCsvSheet.cell(row = row, column = 4).value,\
            vdrCsvSheet.cell(row = row, column = 5).value,\
            vdrCsvSheet.cell(row = row, column = 6).value,\
            vdrCsvSheet.cell(row = row, column = 7).value,\
            vdrCsvSheet.cell(row = row, column = 8).value,\
            vdrCsvSheet.cell(row = row, column = 9).value,\
            vdrCsvSheet.cell(row = row, column = 10).value,\
            vdrCsvSheet.cell(row = row, column = 11).value,\
            vdrCsvSheet.cell(row = row, column = 12).value,\
            vdrCsvSheet.cell(row = row, column = 13).value,\
            vdrCsvSheet.cell(row = row, column = 14).value,\
            vdrCsvSheet.cell(row = row, column = 15).value,\
            vdrCsvSheet.cell(row = row, column = 16).value,\
            vdrCsvSheet.cell(row = row, column = 17).value,\
            vdrCsvSheet.cell(row = row, column = 18).value\
            ] for row in vdrRows}


os.chdir(pathToDocuments)
folders = os.listdir()

def getFormat(page):
  box = page.mediaBox
  upperRight = box.upperRight
  shortSide = min(upperRight)
  longSide = max(upperRight)
  if shortSide < 600.0 and longSide < 900.0: return 'A4'
  elif shortSide < 900.0 and longSide < 1200.0: return 'A3'
  elif shortSide < 1200.0 and longSide < 1700.0: return 'A2'
  elif shortSide < 1700.0 and longSide < 2400.0: return 'A1'
  elif shortSide < 2400.0 and longSide < 3400.0: return 'A0'

issueCSVDict = {'IFI':'IFI - Выпущено для информации',\
                'IFR':'IFR - Выпущено для рассмотрения',\
                'AFC':'AFC - Утверждено для строительства',\
                'IFA':'IFA - Выпущено для согласования',\
                'IFC':'IFC - Выпущено для строительства',\
                'IFD':'IFD - Выпущено для проектирования',\
                'IFH':'IFH - Выпущено для HAZOP',\
                'IFP':'IFP - Выпущено для закупки',\
                'IFQ':'IFQ - Выпущено для предложения цены',\
                'IFU':'IFU - Выпущено для использования',\
                'SD':'SD - Заменен',\
                'VD':'VD - Аннулирован'}

transData = []
transData_CSV = []

for folder in folders:
  pathToFolder = pathToDocuments + '\\'+ folder
  os.chdir(pathToFolder)
  files = os.listdir()
  pdfs = list(filter(lambda f: '.pdf' in f, files))
  for pd in pdfs:
    language = pd.split('_')[-1].split('.')[0]
    packageCode = pd.split('-')[4]

    pathToPdf = pathToFolder + '\\' + pd
    pdf = PdfFileReader(open(pathToPdf,'rb'))
    numberOfSheets = pdf.getNumPages()
    formats = []
    for n in range(numberOfSheets):
      page = pdf.getPage(n)
      formatPage = getFormat(page)
      if formatPage not in formats:
        formats.append(formatPage)
    sheetFormat = '/'.join(formats)
    dwg = pd.replace('.pdf', '.dwg')


    try:
      vdrRow = vdrDict[pd]
      vdrCsvRow = vdrCsvDict[pd]

      tdRow = [vdrRow[8],\
              None,\
              None,\
              vdrRow[3],\
              language,\
              vdrRow[5],\
              vdrRow[4],\
              vdrRow[1],\
              packageCode,\
              vdrRow[6],\
              vdrRow[9],\
              vdrRow[0],\
              vdrRow[7],\
              vdrRow[2],\
              numberOfSheets,\
              sheetFormat,\
              'pdf',\
              pd]
      tdRow_csv =[None,\
              vdrRow[3],\
              None,\
              vdrRow[5],\
              vdrRow[4],\
              vdrRow[2],\
              vdrRow[2],\
              vdrRow[2],\
              vdrRow[9],\
              vdrRow[7],\
              issueCSVDict[vdrRow[6]],\
              'CPECC',\
              '0055',\
              None,\
              vdrRow[8],\
              '4 - ГПЗ',\
              '4.1',\
              vdrCsvRow[0],\
              vdrCsvRow[1],\
              vdrCsvRow[2],\
              vdrCsvRow[3],\
              vdrCsvRow[4],\
              vdrCsvRow[5],\
              vdrCsvRow[6],\
              vdrCsvRow[7],\
              vdrCsvRow[8],\
              vdrCsvRow[9],\
              vdrCsvRow[10],\
              vdrCsvRow[11],\
              vdrCsvRow[12],\
              vdrCsvRow[13],\
              vdrCsvRow[14],\
              vdrCsvRow[15],\
              vdrCsvRow[16],\
              numberOfSheets,\
              1,\
              None,\
              None,\
              None,\
              None,\
              None,\
              language,\
              sheetFormat,\
              None,\
              number,\
              None,\
              'Latest',\
              None,\
              pd,\
              'Native Format',\
              dwg]
    except:
      tdRow = [None,\
              None, \
              None, \
              None, \
              language, \
              None, \
              None, \
              None, \
              packageCode, \
              None, \
              None, \
              None, \
              None, \
              None, \
              numberOfSheets, \
              sheetFormat, \
              'pdf', \
              pd]
      tdRow_csv =[None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              'CPECC',\
              '0055',\
              None,\
              None,\
              '4 - ГПЗ',\
              '4.1',\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              None,\
              numberOfSheets,\
              1,\
              None,\
              None,\
              None,\
              None,\
              None,\
              language,\
              sheetFormat,\
              None,\
              number,\
              None,\
              'Latest',\
              None,\
              pd,\
              'Native Format',\
              dwg]
    transData.append(tdRow)
    transData_CSV.append(tdRow_csv)


r = 8
for i in transData:
  for k, ii in enumerate(i):
	  templateSheet.cell(row = r, column = k + 1).value = ii
  r += 1

r = 2
for i in transData_CSV:
  for k, ii in enumerate(i):
	  templateSheet_CSV.cell(row = r, column = k + 1).value = ii
  r += 1

os.chdir(pathToRelease)
wbTemplate.save(number + '.xlsx')
wbTemplate_CSV.save(number + '_CSV.xlsx')



